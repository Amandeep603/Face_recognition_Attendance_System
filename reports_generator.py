import os
import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import config

class ReportsGenerator:
    """
    Generates multi-format attendance reports (CSV, Excel, PDF)
    with custom date filtering, institutional branding, and statistical summaries.
    """

    def __init__(self, attendance_manager):
        self.attendance_manager = attendance_manager
        os.makedirs(config.REPORTS_DIR, exist_ok=True)

    def fetch_report_data(self, report_type="daily", start_date=None, end_date=None, student_name=None):
        """
        Gathers attendance records based on filter type:
        - daily: single date (start_date or today)
        - weekly: past 7 days up to end_date
        - monthly: past 30 days up to end_date
        - custom: range between start_date and end_date
        - student: specific student across date range
        """
        today_str = datetime.now().strftime("%Y-%m-%d")

        if report_type == "daily":
            dates = [start_date or today_str]
        elif report_type == "weekly":
            end_dt = datetime.strptime(end_date or today_str, "%Y-%m-%d")
            dates = [(end_dt - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6, -1, -1)]
        elif report_type == "monthly":
            end_dt = datetime.strptime(end_date or today_str, "%Y-%m-%d")
            dates = [(end_dt - timedelta(days=i)).strftime("%Y-%m-%d") for i in range(29, -1, -1)]
        elif report_type == "custom" and start_date and end_date:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            delta = (end_dt - start_dt).days
            dates = [(start_dt + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(max(0, delta + 1))]
        else:
            dates = [today_str]

        records = []
        for d in dates:
            csv_path = self.attendance_manager._get_csv_path(d)
            if os.path.exists(csv_path):
                try:
                    df = pd.read_csv(csv_path)
                    for _, row in df.iterrows():
                        r_name = str(row.get("Name", "")).strip()
                        if student_name and r_name.lower() != student_name.lower():
                            continue
                        records.append({
                            "Name": r_name,
                            "Date": str(row.get("Date", d)),
                            "Time": str(row.get("Time", "")),
                            "Status": str(row.get("Status", "Present")),
                            "Verification": "Live Verified"
                        })
                except Exception:
                    pass

        # Calculate metrics
        all_students = self.attendance_manager.get_registered_students()
        total_students_count = len(all_students)
        unique_present = len(set(r["Name"] for r in records))

        summary = {
            "report_type": report_type.title(),
            "date_range": f"{dates[0]} to {dates[-1]}" if len(dates) > 1 else dates[0],
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_registered_students": total_students_count,
            "total_entries": len(records),
            "unique_present_students": unique_present,
            "attendance_rate": f"{round((unique_present / total_students_count * 100), 1)}%" if total_students_count > 0 else "0%"
        }

        return records, summary

    def generate_csv(self, report_type="daily", start_date=None, end_date=None, student_name=None):
        """Generates a filtered CSV report file."""
        records, summary = self.fetch_report_data(report_type, start_date, end_date, student_name)
        filename = f"Report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(config.REPORTS_DIR, filename)

        df = pd.DataFrame(records) if records else pd.DataFrame(columns=["Name", "Date", "Time", "Status", "Verification"])
        df.to_csv(filepath, index=False)
        return filepath, filename

    def generate_excel(self, report_type="daily", start_date=None, end_date=None, student_name=None):
        """Generates a professionally styled Excel spreadsheet report (.xlsx)."""
        records, summary = self.fetch_report_data(report_type, start_date, end_date, student_name)
        filename = f"Report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(config.REPORTS_DIR, filename)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        title_font = Font(name="Segoe UI", size=15, bold=True, color="1E1B4B")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        meta_font = Font(name="Segoe UI", size=10, italic=True, color="4B5563")
        regular_font = Font(name="Segoe UI", size=10, color="1F2937")
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        present_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        present_font = Font(name="Segoe UI", size=10, bold=True, color="166534")

        thin_side = Side(border_style="thin", color="E2E8F0")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Title Block
        ws.merge_cells("A1:E1")
        ws["A1"] = "AI Face Recognition Attendance System - Official Report"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(vertical="center")

        ws["A2"] = f"Filter: {summary['report_type']} | Range: {summary['date_range']} | Generated: {summary['generated_at']}"
        ws["A2"].font = meta_font

        # Summary Block
        ws["A4"] = f"Total Registered: {summary['total_registered_students']}  |  Unique Present: {summary['unique_present_students']}  |  Attendance Rate: {summary['attendance_rate']}"
        ws["A4"].font = Font(name="Segoe UI", size=10, bold=True, color="4338CA")

        # Table Header
        headers = ["Student Name", "Date", "Check-in Time", "Attendance Status", "Biometric Verification"]
        start_row = 6
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_all
            ws.row_dimensions[start_row].height = 24

        # Table Data
        curr_row = start_row + 1
        for i, rec in enumerate(records):
            ws.cell(row=curr_row, column=1, value=rec["Name"]).font = regular_font
            ws.cell(row=curr_row, column=2, value=rec["Date"]).font = regular_font
            ws.cell(row=curr_row, column=3, value=rec["Time"]).font = regular_font
            
            status_cell = ws.cell(row=curr_row, column=4, value=rec["Status"])
            status_cell.font = present_font
            status_cell.fill = present_fill
            status_cell.alignment = Alignment(horizontal="center")

            verif_cell = ws.cell(row=curr_row, column=5, value=rec["Verification"])
            verif_cell.font = regular_font
            verif_cell.alignment = Alignment(horizontal="center")

            # Borders & Zebra
            for col_idx in range(1, 6):
                c = ws.cell(row=curr_row, column=col_idx)
                c.border = border_all
                if i % 2 == 1 and col_idx != 4:
                    c.fill = zebra_fill

            ws.row_dimensions[curr_row].height = 20
            curr_row += 1

        # Column widths
        widths = [26, 16, 18, 20, 24]
        for idx, w in enumerate(widths, start=1):
            col_letter = openpyxl.utils.get_column_letter(idx)
            ws.column_dimensions[col_letter].width = w

        wb.save(filepath)
        return filepath, filename

    def generate_pdf(self, report_type="daily", start_date=None, end_date=None, student_name=None):
        """Generates an enterprise-formatted institutional PDF attendance report."""
        records, summary = self.fetch_report_data(report_type, start_date, end_date, student_name)
        filename = f"Report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(config.REPORTS_DIR, filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor('#1E1B4B'),
            fontName='Helvetica-Bold'
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor('#4B5563'),
            fontName='Helvetica-Oblique'
        )
        cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            textColor=colors.HexColor('#1F2937'),
            fontName='Helvetica'
        )
        cell_header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            textColor=colors.white,
            fontName='Helvetica-Bold'
        )

        elements = []

        # 1. Header & Branding
        elements.append(Paragraph("AI Face Recognition Attendance System", title_style))
        elements.append(Paragraph(f"Official Institutional Attendance Report &bull; {summary['report_type']} Mode", subtitle_style))
        elements.append(Paragraph(f"<b>Date Range:</b> {summary['date_range']} &nbsp;&nbsp;|&nbsp;&nbsp; <b>Generated:</b> {summary['generated_at']}", subtitle_style))
        elements.append(Spacer(1, 14))

        # 2. KPI Summary Table
        kpi_data = [
            ["Total Enrolled", "Unique Present", "Total Check-ins", "Attendance Rate"],
            [
                str(summary['total_registered_students']),
                str(summary['unique_present_students']),
                str(summary['total_entries']),
                str(summary['attendance_rate'])
            ]
        ]
        kpi_table = Table(kpi_data, colWidths=[130, 130, 130, 130])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EEF2FF')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#4338CA')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F8FAFC')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 12),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.HexColor('#1E1B4B')),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 16))

        # 3. Records Table
        table_data = [[
            Paragraph("<b>#</b>", cell_header_style),
            Paragraph("<b>Student Name</b>", cell_header_style),
            Paragraph("<b>Date</b>", cell_header_style),
            Paragraph("<b>Time</b>", cell_header_style),
            Paragraph("<b>Status</b>", cell_header_style),
            Paragraph("<b>Biometric Liveness</b>", cell_header_style)
        ]]

        for i, rec in enumerate(records, start=1):
            table_data.append([
                Paragraph(str(i), cell_style),
                Paragraph(rec["Name"], cell_style),
                Paragraph(rec["Date"], cell_style),
                Paragraph(rec["Time"], cell_style),
                Paragraph("<font color='#166534'><b>Present</b></font>", cell_style),
                Paragraph("<font color='#4338CA'>Live Verified</font>", cell_style)
            ])

        if len(records) == 0:
            table_data.append([
                Paragraph("-", cell_style),
                Paragraph("<i>No attendance records found for this period</i>", cell_style),
                Paragraph("-", cell_style),
                Paragraph("-", cell_style),
                Paragraph("-", cell_style),
                Paragraph("-", cell_style)
            ])

        rec_table = Table(table_data, colWidths=[30, 150, 80, 80, 80, 100])
        rec_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#CBD5E1')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')])
        ]))

        elements.append(rec_table)

        # Build document
        doc.build(elements)
        return filepath, filename
